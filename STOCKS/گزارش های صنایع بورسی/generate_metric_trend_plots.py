from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Iterable

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
except Exception:
    arabic_reshaper = None
    get_display = None


def sanitize_name(text: str) -> str:
    text = re.sub(r'[\\/*?:"<>|]', '_', str(text))
    text = re.sub(r'\s+', ' ', text).strip()
    return text or 'unknown'


def rtl_text(text: str) -> str:
    s = str(text)
    if arabic_reshaper is not None and get_display is not None:
        try:
            return get_display(arabic_reshaper.reshape(s))
        except Exception:
            pass
    # Fallback: reverse to improve readability in LTR rendering backends.
    return s[::-1]


def to_english_digits(text: str) -> str:
    return str(text).translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789"))


def parse_year(value_header: str) -> int | None:
    if not value_header:
        return None
    header = to_english_digits(str(value_header))
    m = re.search(r"(\d{4})", header)
    return int(m.group(1)) if m else None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s == "-":
        return None
    s = to_english_digits(s).replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def discover_modified_workbooks(data_root: Path, industry_filter: str | None, report_filter: str | None) -> list[tuple[Path, str, str]]:
    out: list[tuple[Path, str, str]] = []
    for modified_dir in sorted(p for p in data_root.rglob('*') if p.is_dir() and p.name.endswith(' modified')):
        industry = modified_dir.parent.name
        report_folder = modified_dir.name[:-len(' modified')]

        if industry_filter and industry != industry_filter:
            continue
        if report_filter and report_folder != report_filter:
            continue

        for wb_path in sorted(modified_dir.glob('*.xlsx')) + sorted(modified_dir.glob('*.xlsm')):
            if wb_path.name.startswith('~$'):
                continue
            if wb_path.stem.endswith(' - report'):
                continue
            out.append((wb_path, industry, report_folder))
    return out


def get_metric_blocks(ws) -> list[tuple[str, list[tuple[int, int]]]]:
    # row 4: metric names (merged blocks start), row 5: year headers (Value/YoY)
    max_col = ws.max_column
    starts = []
    for c in range(2, max_col + 1):
        v = ws.cell(4, c).value
        if v is not None and str(v).strip() != "":
            starts.append(c)

    blocks: list[tuple[str, list[tuple[int, int]]]] = []
    for i, sc in enumerate(starts):
        ec = (starts[i + 1] - 1) if i + 1 < len(starts) else max_col
        metric = str(ws.cell(4, sc).value).strip()

        year_cols: list[tuple[int, int]] = []
        for c in range(sc, ec + 1):
            h = ws.cell(5, c).value
            hs = str(h).strip() if h is not None else ""
            if "Value" not in hs:
                continue
            year = parse_year(hs)
            if year is not None:
                year_cols.append((year, c))

        year_cols.sort(key=lambda x: x[0])
        if year_cols:
            blocks.append((metric, year_cols))
    return blocks


def iter_symbol_rows(ws, start_row: int = 6) -> Iterable[tuple[str, int]]:
    r = start_row
    while r <= ws.max_row:
        symbol = ws.cell(r, 1).value
        if symbol is None or str(symbol).strip() == "":
            break
        yield str(symbol).strip(), r
        r += 1


def ratio_to_compressed(ratio: float) -> float:
    # Compress very large multipliers while preserving direction and ordering.
    delta = ratio - 1.0
    return math.copysign(math.log1p(abs(delta)), delta)


def compressed_to_ratio(value: float) -> float:
    delta = math.copysign(math.expm1(abs(value)), value)
    return 1.0 + delta


def plot_metric(
    metric: str,
    year_cols: list[tuple[int, int]],
    ws,
    title_prefix: str,
    out_png: Path,
    scale_mode: str = "compressed",
):
    years_all = [y for y, _ in year_cols]

    plt.rcParams['font.family'] = ['Arial Unicode MS', 'Arial', 'DejaVu Sans']
    fig, ax = plt.subplots(figsize=(14, 8))

    plotted = 0
    y_all = []
    for symbol, row in iter_symbol_rows(ws):
        xs = []
        ys = []
        for y, c in year_cols:
            v = to_float(ws.cell(row, c).value)
            if v is None:
                continue
            xs.append(y)
            ys.append(v)

        if not xs:
            continue

        # Trend-only comparison: normalize each series by its first available value.
        first_v = ys[0]
        if first_v is None or abs(first_v) < 1e-12:
            continue
        ys_norm = [v / first_v for v in ys]
        if scale_mode == "compressed":
            ys_plot = [ratio_to_compressed(v) for v in ys_norm]
        else:
            ys_plot = ys_norm

        if len(xs) == 1:
            ax.scatter(xs, ys_plot, s=24)
        else:
            ax.plot(xs, ys_plot, marker='o', linewidth=1.6, markersize=3)

        # Label each line (or point) on the right side at the last available point.
        ax.annotate(rtl_text(symbol), (xs[-1], ys_plot[-1]), xytext=(6, 0), textcoords='offset points', fontsize=8, va='center')
        y_all.extend(ys_plot)
        plotted += 1

    if plotted == 0:
        plt.close(fig)
        return False

    ax.set_title(rtl_text(f"{title_prefix} | {metric}"), fontsize=12)
    ax.set_xlabel("Year")
    if scale_mode == "compressed":
        ax.set_ylabel("Trend Multiplier (Compressed Scale, Base=1.0)")
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: f"{compressed_to_ratio(x):.2f}x"))
    else:
        ax.set_ylabel("Normalized Trend (Base=1.0)")
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: f"{x:.2f}x"))

    if y_all:
        ymin = min(y_all)
        ymax = max(y_all)
        if math.isfinite(ymin) and math.isfinite(ymax):
            pad = (ymax - ymin) * 0.08
            if pad <= 0:
                pad = 0.1
            ax.set_ylim(ymin - pad, ymax + pad)

    ax.set_xticks(sorted(set(years_all)))
    ax.grid(True, alpha=0.25)

    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=170)
    plt.close(fig)
    return True


def process_workbook(wb_path: Path, industry: str, report_folder: str, overwrite: bool, scale_mode: str) -> tuple[int, int]:
    wb = load_workbook(wb_path, data_only=True)
    ws = wb['Normalized'] if 'Normalized' in wb.sheetnames else wb[wb.sheetnames[0]]

    report_out_dir = wb_path.parent / sanitize_name(wb_path.stem)
    blocks = get_metric_blocks(ws)

    created = 0
    skipped = 0
    for metric, year_cols in blocks:
        png_name = f"{sanitize_name(metric)}.png"
        out_png = report_out_dir / png_name
        if out_png.exists() and not overwrite:
            skipped += 1
            continue

        ok = plot_metric(
            metric=metric,
            year_cols=year_cols,
            ws=ws,
            title_prefix=f"{industry} - {report_folder}",
            out_png=out_png,
            scale_mode=scale_mode,
        )
        if ok:
            created += 1
        else:
            skipped += 1

    return created, skipped


def main():
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Generate metric trend plots from all modified report workbooks")
    parser.add_argument('--data-root', default=str(script_dir / 'data'), help='Root data folder')
    parser.add_argument('--industry', default=None, help='Optional industry folder filter')
    parser.add_argument('--report-folder', default=None, help='Optional report folder filter (e.g. سود و زیان)')
    parser.add_argument('--skip-existing-images', action='store_true', help='Skip PNG files that already exist')
    parser.add_argument(
        '--scale-mode',
        default='compressed',
        choices=['compressed', 'raw'],
        help='Y-axis scale mode: compressed keeps all trends visible while preserving upside ordering',
    )
    args = parser.parse_args()

    data_root = Path(args.data_root).resolve()
    if not data_root.exists():
        raise FileNotFoundError(f"Data root not found: {data_root}")

    jobs = discover_modified_workbooks(data_root, args.industry, args.report_folder)
    if not jobs:
        print(f"No modified report workbooks found under: {data_root}")
        return

    print(f"Discovered {len(jobs)} modified workbook(s)")
    total_created = 0
    total_skipped = 0
    total_failed = 0

    for wb_path, industry, report_folder in jobs:
        try:
            created, skipped = process_workbook(
                wb_path=wb_path,
                industry=industry,
                report_folder=report_folder,
                overwrite=not args.skip_existing_images,
                scale_mode=args.scale_mode,
            )
            total_created += created
            total_skipped += skipped
            print(f"[ok] {wb_path.name} -> folder '{wb_path.stem}' | images created={created}, skipped={skipped}")
        except Exception as e:
            total_failed += 1
            print(f"[error] {wb_path}: {e}")

    print(f"Done. images_created={total_created}, images_skipped={total_skipped}, workbooks_failed={total_failed}")


if __name__ == '__main__':
    main()
