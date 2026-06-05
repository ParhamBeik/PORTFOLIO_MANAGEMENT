from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_KEYWORDS = {"سرفصل", "سال مالی", "نوع گزارش", "نوع داده", "تجدید ارائه", "حسابرسی"}

# --------------------------- USER CONFIG ---------------------------
# Optional filters (None means process all industries/report folders)
TARGET_INDUSTRY_FOLDER: str | None = None
TARGET_REPORT_FOLDER: str | None = None
# Input values are in billion Rials; divide by 10 to show billion Tomans.
VALUE_SCALE_DIVISOR = 10
VALUE_UNIT_LABEL = "ارقام: میلیارد تومان (تبدیل‌شده از میلیارد ریال)"
# Drop a metric if this share of value-cells is empty.
# You can use ratio (0.80) or percent (80).
EMPTY_METRIC_DROP_THRESHOLD = 0.8
# For pruning only: treat 0 values as empty (common in sparse statements).
TREAT_ZERO_AS_EMPTY_FOR_PRUNING = True
# -------------------------------------------------------------------

METRIC_ORDER = [
    "درآمد حاصل از خدمات و فروش",
    "هزینه خدمات و فروش پیمانکاری\u200cها",
    "سود حاصل از سایر فعالیت\u200cها",
    "جمع درآمدها",
    "بهای تمام شده کالای فروش رفته",
    "سود ناویژه",
    "هزینه\u200cهای عمومی و اداری",
    "خالص سایر درآمدها و هزینه\u200cهای عملیاتی",
    "هزینه مطالبات مشکوک الوصول",
    "هزینه\u200cهای مالی",
    "هزینه (درآمد) عملیاتی",
    "سایر هزینه\u200cهای مالی",
    "درآمد حاصل از سرمایه گذاری",
    "خالص سایر درآمدها (هزینه\u200cها)",
    "سهم گروه از سود شرکت\u200cهای وابسته",
    "سهم از سود (زیان) خالص انتهای سرمایه\u200cگذاری در اصول و رویه\u200cهای حسابداری و مالیات",
    "سود قبل از کسر مالیات",
    "مالیات",
    "مالیات سال قبل",
    "سود (زیان) ناخالص سال جاری",
    "سود (زیان) ویژه پس از کسر مالیات",
    "سود قابل تخصیص",
    "افزایش سرمایه (سود زیان) انباشته",
    "سود و زیان انباشته در پایان دوره",
    "سرمایه",
    "EPS خالص",
]

# If these metrics increase, that is usually negative from an analysis perspective.
BAD_IF_UP = {
    "بهای تمام شده کالای فروش رفته",
    "هزینه\u200cهای مالی",
    "مالیات",
}


@dataclass(frozen=True)
class JalaliDate:
    year: int
    month: int
    day: int


def sanitize_name(text: str) -> str:
    text = re.sub(r"[\\/*?:\"<>|]", "_", str(text))
    text = re.sub(r"\s+", "_", text).strip("_")
    return text or "unknown"


def to_persian_digits(text: str) -> str:
    return str(text).translate(str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹"))


def parse_jalali_date(value: object) -> JalaliDate | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None
    trans = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
    s = s.translate(trans)
    m = re.fullmatch(r"(\d{4})\/(\d{1,2})\/(\d{1,2})", s)
    if not m:
        return None
    y, mo, d = map(int, m.groups())
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    return JalaliDate(y, mo, d)


def to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None
    s = s.translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789"))
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def first_non_empty(values: Iterable[object], default: str) -> str:
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() != "nan":
            return s
    return default


def find_data_start(df_raw: pd.DataFrame) -> int:
    for r in range(len(df_raw)):
        first = str(df_raw.iat[r, 0]).strip() if df_raw.iat[r, 0] is not None else ""
        if not first or first.lower() == "nan":
            continue
        if first in HEADER_KEYWORDS:
            continue
        return r
    raise ValueError("Could not find start of company rows")


def extract_column_meta(df_raw: pd.DataFrame):
    metric_by_col: dict[int, str] = {}
    date_by_col: dict[int, JalaliDate | None] = {}
    year_by_col: dict[int, int | None] = {}
    mmdd_by_col: dict[int, tuple[int, int] | None] = {}

    current_metric = None
    for c in range(1, df_raw.shape[1]):
        mval = df_raw.iat[0, c]
        if mval is not None:
            ms = str(mval).strip()
            if ms and ms.lower() != "nan":
                current_metric = ms
        metric_by_col[c] = current_metric

        d = parse_jalali_date(df_raw.iat[1, c])
        date_by_col[c] = d
        year_by_col[c] = d.year if d else None
        mmdd_by_col[c] = (d.month, d.day) if d else None

    valid_cols = [
        c for c in range(1, df_raw.shape[1])
        if metric_by_col.get(c) and date_by_col.get(c)
    ]
    return metric_by_col, date_by_col, year_by_col, mmdd_by_col, valid_cols


def determine_anchor_for_symbol(row: pd.Series, valid_cols: list[int], mmdd_by_col: dict[int, tuple[int, int] | None]) -> tuple[int, int] | None:
    counts: Counter[tuple[int, int]] = Counter()
    for c in valid_cols:
        if to_float(row.iat[c]) is not None:
            mmdd = mmdd_by_col.get(c)
            if mmdd:
                counts[mmdd] += 1
    if not counts:
        return None
    return max(counts.items(), key=lambda x: x[1])[0]


def mmdd_distance(a: tuple[int, int], b: tuple[int, int]) -> int:
    # Approximate distance in calendar (month/day only).
    return abs((a[0] * 31 + a[1]) - (b[0] * 31 + b[1]))


def choose_best_column(
    row: pd.Series,
    candidate_cols: list[int],
    anchor_mmdd: tuple[int, int] | None,
    mmdd_by_col: dict[int, tuple[int, int] | None],
    date_by_col: dict[int, JalaliDate | None],
) -> int | None:
    present = [c for c in candidate_cols if to_float(row.iat[c]) is not None]
    if not present:
        return None
    if anchor_mmdd:
        exact = [c for c in present if mmdd_by_col.get(c) == anchor_mmdd]
        if exact:
            return max(exact, key=lambda c: (date_by_col[c].month, date_by_col[c].day))

        nearest = min(
            present,
            key=lambda c: mmdd_distance(mmdd_by_col.get(c) or (1, 1), anchor_mmdd),
        )
        return nearest

    return max(present, key=lambda c: (date_by_col[c].month, date_by_col[c].day))


def transform(df_raw: pd.DataFrame):
    metric_by_col, date_by_col, year_by_col, mmdd_by_col, valid_cols = extract_column_meta(df_raw)
    start_row = find_data_start(df_raw)

    years = sorted({year_by_col[c] for c in valid_cols if year_by_col[c] is not None})
    # Keep source metric order for non-priority metrics.
    metrics_in_source_order: list[str] = []
    for c in valid_cols:
        m = metric_by_col[c]
        if m and m not in metrics_in_source_order:
            metrics_in_source_order.append(m)
    ordered_metrics = [m for m in METRIC_ORDER if m in metrics_in_source_order] + [
        m for m in metrics_in_source_order if m not in METRIC_ORDER
    ]

    values_rows: list[dict] = []
    yoy_rows: list[dict] = []

    cols_by_metric_year: dict[tuple[str, int], list[int]] = defaultdict(list)
    for c in valid_cols:
        metric = metric_by_col[c]
        year = year_by_col[c]
        if metric and year:
            cols_by_metric_year[(metric, year)].append(c)

    for r in range(start_row, len(df_raw)):
        symbol = first_non_empty([df_raw.iat[r, 0]], "")
        if not symbol or symbol in HEADER_KEYWORDS:
            continue

        row = df_raw.iloc[r]
        anchor = determine_anchor_for_symbol(row, valid_cols, mmdd_by_col)

        val_record = {"Symbol": symbol}
        yoy_record = {"Symbol": symbol}

        for metric in ordered_metrics:
            prev_value = None
            for year in years:
                cols = cols_by_metric_year.get((metric, year), [])
                chosen = choose_best_column(row, cols, anchor, mmdd_by_col, date_by_col)
                value = to_float(row.iat[chosen]) if chosen else None
                val_record[(metric, year)] = value

                pct = None
                if prev_value is not None and value is not None and prev_value != 0:
                    pct = (value - prev_value) / abs(prev_value)
                yoy_record[(metric, year)] = pct

                if value is not None:
                    prev_value = value

        values_rows.append(val_record)
        yoy_rows.append(yoy_record)

    df_values = pd.DataFrame(values_rows).set_index("Symbol")
    df_yoy = pd.DataFrame(yoy_rows).set_index("Symbol")

    df_values.columns = pd.MultiIndex.from_tuples(df_values.columns, names=["Metric", "Year"])
    df_yoy.columns = pd.MultiIndex.from_tuples(df_yoy.columns, names=["Metric", "Year"])

    return df_values, df_yoy, ordered_metrics, years


def prune_sparse_metrics(
    df_values: pd.DataFrame,
    ordered_metrics: list[str],
    years: list[int],
    empty_threshold: float,
):
    threshold = float(empty_threshold)
    if threshold > 1:
        threshold = threshold / 100.0
    threshold = max(0.0, min(1.0, threshold))

    kept_metrics: list[str] = []
    for metric in ordered_metrics:
        metric_cols = [(metric, y) for y in years if (metric, y) in df_values.columns]
        if not metric_cols:
            continue
        total_cells = len(df_values.index) * len(metric_cols)
        if total_cells == 0:
            continue
        metric_block = df_values[metric_cols]
        if TREAT_ZERO_AS_EMPTY_FOR_PRUNING:
            metric_num = metric_block.apply(pd.to_numeric, errors="coerce")
            non_empty_mask = metric_num.notna() & (metric_num.abs() > 1e-12)
            non_empty = int(non_empty_mask.sum().sum())
        else:
            non_empty = int(metric_block.notna().sum().sum())
        empty_ratio = 1 - (non_empty / total_cells)
        if empty_ratio < threshold:
            kept_metrics.append(metric)

    keep_cols = [(m, y) for m in kept_metrics for y in years if (m, y) in df_values.columns]
    return df_values[keep_cols], kept_metrics


def write_output(
    output_path: Path,
    df_values: pd.DataFrame,
    df_yoy: pd.DataFrame,
    ordered_metrics: list[str],
    years: list[int],
    report_type: str,
    data_type: str,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Normalized"

    ws["A1"] = "Normalized Financial Report (Fiscal-Offset Aware)"
    ws["A1"].font = Font(bold=True, size=13, color="1F2937")
    ws["A2"] = f"Report Type: {report_type} | Data Type: {data_type}"
    ws["A2"].font = Font(color="4B5563")
    ws["A3"] = VALUE_UNIT_LABEL
    ws["A3"].font = Font(color="1F2937", bold=True)

    start_header_row = 4
    ws.cell(start_header_row, 1, "نماد")
    ws.cell(start_header_row + 1, 1, "Symbol")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9E1F2")

    for rr in (start_header_row, start_header_row + 1):
        c = ws.cell(rr, 1)
        c.fill = header_fill if rr == start_header_row else subheader_fill
        c.font = Font(color="FFFFFF" if rr == start_header_row else "1F2937", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    col = 2
    layout: list[tuple[str, int, int, int]] = []  # metric, year, value_col, yoy_col

    for metric in ordered_metrics:
        metric_start_col = col
        for year in years:
            ws.cell(start_header_row + 1, col, f"{year} Value")
            ws.cell(start_header_row + 1, col + 1, f"{year} YoY%")

            for cc in (col, col + 1):
                sub = ws.cell(start_header_row + 1, cc)
                sub.fill = subheader_fill
                sub.font = Font(color="1F2937", bold=True)
                sub.alignment = Alignment(horizontal="center", vertical="center")

            layout.append((metric, year, col, col + 1))
            col += 2
        metric_end_col = col - 1
        ws.merge_cells(
            start_row=start_header_row,
            start_column=metric_start_col,
            end_row=start_header_row,
            end_column=metric_end_col,
        )
        metric_cell = ws.cell(start_header_row, metric_start_col, metric)
        metric_cell.fill = header_fill
        metric_cell.font = Font(color="FFFFFF", bold=True)
        metric_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = start_header_row + 2

    good_fill_light = PatternFill("solid", fgColor="E2F0D9")
    good_fill_dark = PatternFill("solid", fgColor="92D050")
    bad_fill_light = PatternFill("solid", fgColor="FCE4D6")
    bad_fill_dark = PatternFill("solid", fgColor="FF6666")
    na_fill = PatternFill("solid", fgColor="F2F2F2")

    for i, symbol in enumerate(df_values.index, start=data_start):
        ws.cell(i, 1, symbol)
        ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 1).alignment = Alignment(horizontal="center", vertical="center")

        for metric, year, value_col, yoy_col in layout:
            v = df_values.at[symbol, (metric, year)] if (metric, year) in df_values.columns else None
            y = df_yoy.at[symbol, (metric, year)] if (metric, year) in df_yoy.columns else None
            if pd.isna(v):
                v = None
            if pd.isna(y):
                y = None
            if v is not None:
                v = v / VALUE_SCALE_DIVISOR

            vc = ws.cell(i, value_col, v)
            yc = ws.cell(i, yoy_col, y)

            vc.number_format = '#,##0_);(#,##0);"-"'
            yc.number_format = '0.0%;-0.0%;"-"'

            vc.alignment = Alignment(horizontal="center", vertical="center")
            yc.alignment = Alignment(horizontal="center", vertical="center")
            vc.font = Font(color="000000")
            yc.font = Font(color="000000")

            if v is None:
                vc.value = "-"
                vc.fill = na_fill
                vc.alignment = Alignment(horizontal="center", vertical="center")

            if y is None:
                yc.value = "-"
                yc.fill = na_fill
                yc.alignment = Alignment(horizontal="center", vertical="center")
            else:
                is_good = y < 0 if metric in BAD_IF_UP else y > 0
                strong = abs(y) >= 0.30
                if is_good:
                    yc.fill = good_fill_dark if strong else good_fill_light
                else:
                    yc.fill = bad_fill_dark if strong else bad_fill_light

    # Draw clear borders around each metric block.
    thin = Side(style="thin", color="BFBFBF")
    medium = Side(style="medium", color="404040")
    end_data_row = data_start + len(df_values.index) - 1
    metric_blocks: dict[str, tuple[int, int]] = {}
    for metric, _, value_col, yoy_col in layout:
        if metric not in metric_blocks:
            metric_blocks[metric] = (value_col, yoy_col)
        else:
            metric_blocks[metric] = (metric_blocks[metric][0], yoy_col)

    for _, (start_col, end_col) in metric_blocks.items():
        for rr in range(start_header_row, end_data_row + 1):
            for cc in range(start_col, end_col + 1):
                left = medium if cc == start_col else thin
                right = medium if cc == end_col else thin
                top = medium if rr == start_header_row else thin
                bottom = medium if rr == end_data_row else thin
                ws.cell(rr, cc).border = Border(left=left, right=right, top=top, bottom=bottom)

    ws.freeze_panes = "B6"
    end_data_row = data_start + len(df_values.index) - 1
    ws.auto_filter.ref = f"A5:{get_column_letter(col - 1)}{end_data_row}"

    # Better default spacing/readability.
    ws.row_dimensions[start_header_row].height = 24
    ws.row_dimensions[start_header_row + 1].height = 22
    for r in range(data_start, end_data_row + 1):
        ws.row_dimensions[r].height = 20

    # Auto-size columns from actual content length.
    for c in range(1, col):
        max_len = 0
        for r in range(1, end_data_row + 1):
            value = ws.cell(r, c).value
            if value is None:
                continue
            cell_len = len(str(value))
            if cell_len > max_len:
                max_len = cell_len

        # Keep widths bounded for consistency.
        if c == 1:
            width = max(14, min(max_len + 3, 28))
        else:
            width = max(12, min(max_len + 3, 22))
        ws.column_dimensions[get_column_letter(c)].width = width

    legend_col = col + 1
    ws.cell(4, legend_col, "Legend").font = Font(bold=True)
    ws.cell(5, legend_col, "Green: favorable trend")
    ws.cell(6, legend_col, "Red: unfavorable trend")
    ws.cell(7, legend_col, "- : no data")
    ws.cell(8, legend_col, "Cost metrics are inverse-colored")

    wb.save(output_path)


def process_one_file(
    input_file: Path,
    modified_dir: Path,
    industry_name: str,
    report_folder_name: str,
    overwrite_existing: bool,
) -> tuple[Path, str]:
    df_raw = pd.read_excel(input_file, header=None)

    report_type = first_non_empty(df_raw.iloc[2, 1:].tolist(), "UnknownReport") if len(df_raw) > 2 else "UnknownReport"
    data_type = first_non_empty(df_raw.iloc[3, 1:].tolist(), "UnknownType") if len(df_raw) > 3 else "UnknownType"
    report_type = to_persian_digits(report_type)
    data_type = to_persian_digits(data_type)

    df_values, df_yoy, ordered_metrics, years = transform(df_raw)
    df_values, ordered_metrics = prune_sparse_metrics(
        df_values=df_values,
        ordered_metrics=ordered_metrics,
        years=years,
        empty_threshold=EMPTY_METRIC_DROP_THRESHOLD,
    )
    keep_cols = [(m, y) for m in ordered_metrics for y in years if (m, y) in df_yoy.columns]
    df_yoy = df_yoy[keep_cols]

    modified_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"{industry_name} - {report_folder_name} {data_type} {report_type}.xlsx"
    out_name = re.sub(r"[\\/*?:\"<>|]", "_", out_name)
    out_path = modified_dir / out_name
    existed_before = out_path.exists()
    if existed_before and not overwrite_existing:
        return out_path, "skipped"

    write_output(out_path, df_values, df_yoy, ordered_metrics, years, report_type, data_type)
    return out_path, "updated" if existed_before else "created"


def collect_excel_files(folder: Path) -> list[Path]:
    return sorted(
        p for p in folder.iterdir()
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"} and not p.name.startswith("~$")
    )


def discover_jobs(data_root: Path, industry_filter: str | None, report_filter: str | None) -> list[tuple[Path, Path, str, str]]:
    jobs: list[tuple[Path, Path, str, str]] = []
    if not data_root.exists():
        return jobs

    for industry_dir in sorted(p for p in data_root.iterdir() if p.is_dir() and not p.name.startswith(".")):
        industry_name = industry_dir.name
        if industry_filter and industry_name != industry_filter:
            continue

        for report_dir in sorted(p for p in industry_dir.iterdir() if p.is_dir() and not p.name.startswith(".")):
            report_name = report_dir.name
            if report_name.endswith(" modified"):
                continue
            if report_filter and report_name != report_filter:
                continue

            modified_dir = industry_dir / f"{report_name} modified"
            for input_file in collect_excel_files(report_dir):
                jobs.append((input_file, modified_dir, industry_name, report_name))
    return jobs


def main():
    script_dir = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(description="Auto normalize all industry/report Excel files under data/")
    parser.add_argument("--data-root", default=str(script_dir / "data"), help="Root data folder")
    parser.add_argument("--industry", default=TARGET_INDUSTRY_FOLDER, help="Optional industry folder filter")
    parser.add_argument("--report-folder", default=TARGET_REPORT_FOLDER, help="Optional report folder filter (e.g. سود و زیان)")
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        help="Skip outputs that already exist. Default behavior is to overwrite/update existing outputs.",
    )
    args = parser.parse_args()

    data_root = Path(args.data_root).resolve()
    if not data_root.exists():
        raise FileNotFoundError(f"Data root not found: {data_root}")

    overwrite_existing = not args.skip_existing
    jobs = discover_jobs(data_root, args.industry, args.report_folder)
    if not jobs:
        print(f"No eligible Excel files found under: {data_root}")
        return

    print(f"Discovered {len(jobs)} file(s) under: {data_root}")
    created = 0
    updated = 0
    skipped = 0
    failed = 0

    for input_file, modified_dir, industry_name, report_name in jobs:
        try:
            out, status = process_one_file(
                input_file,
                modified_dir,
                industry_name,
                report_name,
                overwrite_existing=overwrite_existing,
            )
            if status == "skipped":
                skipped += 1
                print(f"[skip-exists] {out}")
            else:
                if status == "created":
                    created += 1
                else:
                    updated += 1
                print(f"[ok-{status}] {input_file.name} -> {out.name}")
        except Exception as e:
            failed += 1
            print(f"[error] {input_file}: {e}")

    print(f"Done. created={created}, updated={updated}, skipped={skipped}, failed={failed}")


if __name__ == "__main__":
    main()
