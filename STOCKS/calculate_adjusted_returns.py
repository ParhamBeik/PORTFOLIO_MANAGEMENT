from __future__ import annotations

import argparse
import json
import math
from bisect import bisect_left
from dataclasses import dataclass
from pathlib import Path

import pandas as pd


@dataclass
class StockSeries:
    industry: str
    symbol: str
    dates: list[str]  # ascending YYYY-MM-DD (Jalali)
    closes: list[float]


def parse_adjusted_file(path: Path) -> StockSeries | None:
    try:
        with path.open("r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return None

    cs = payload.get("candlestick_data", {})
    symbol = str(cs.get("l18") or path.parent.parent.parent.name).strip()
    if symbol.endswith("ح"):
        return None
    candles = cs.get("candle_daily_adjusted", [])
    if not isinstance(candles, list) or not candles:
        return None

    # Input is typically newest->oldest; normalize to ascending by date.
    parsed = []
    for c in candles:
        if not isinstance(c, dict):
            continue
        d = c.get("date")
        cl = c.get("close")
        if d is None or cl is None:
            continue
        try:
            d = str(d).strip()
            cl = float(cl)
        except Exception:
            continue
        if not d or cl <= 0:
            continue
        parsed.append((d, cl))

    if not parsed:
        return None

    parsed.sort(key=lambda x: x[0])
    dates = [d for d, _ in parsed]
    closes = [c for _, c in parsed]

    # Expected structure: .../<industry>/<symbol>/Historical/adjusted.json
    industry = path.parents[2].name
    return StockSeries(industry=industry, symbol=symbol, dates=dates, closes=closes)


def find_reference_date(first_dates: list[str], coverage: float = 0.80) -> str:
    if not first_dates:
        raise ValueError("No first dates found")
    first_dates_sorted = sorted(first_dates)
    idx = max(0, math.ceil(coverage * len(first_dates_sorted)) - 1)
    return first_dates_sorted[idx]


def return_metrics(start_price: float, end_price: float) -> tuple[float, float]:
    multiple = end_price / start_price if start_price > 0 else float("nan")
    pct = (multiple - 1.0) * 100 if math.isfinite(multiple) else float("nan")
    return multiple, pct


def build_report_rows(series_list: list[StockSeries], reference_date: str) -> list[dict]:
    rows: list[dict] = []

    for s in series_list:
        first_date = s.dates[0]
        last_date = s.dates[-1]
        first_close = s.closes[0]
        last_close = s.closes[-1]

        total_mult, total_pct = return_metrics(first_close, last_close)

        # If listed before/at reference, start from first point at/after reference date.
        if first_date <= reference_date:
            i = bisect_left(s.dates, reference_date)
            if i >= len(s.dates):
                i = len(s.dates) - 1
            start_date_used = s.dates[i]
            start_close_used = s.closes[i]
            mode = "reference_date"
        else:
            start_date_used = first_date
            start_close_used = first_close
            mode = "new_listing_from_first_date"

        ref_mult, ref_pct = return_metrics(start_close_used, last_close)

        rows.append(
            {
                "industry": s.industry,
                "symbol": s.symbol,
                "first_date": first_date,
                "last_date": last_date,
                "first_close": first_close,
                "last_close": last_close,
                "total_return_multiple": total_mult,
                "total_return_pct": total_pct,
                "reference_date_global": reference_date,
                "start_date_used": start_date_used,
                "start_close_used": start_close_used,
                "return_mode": mode,
                "reference_return_multiple": ref_mult,
                "reference_return_pct": ref_pct,
                "data_points": len(s.dates),
            }
        )

    rows.sort(key=lambda r: (r["reference_return_pct"] if pd.notna(r["reference_return_pct"]) else -1e18), reverse=True)
    return rows


def autosize_excel(path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb["returns"]

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

    # No-decimal numeric presentation with thousands separators.
    # Percent columns are stored as absolute percentages (e.g. 125 means 125%).
    int_cols = [
        "first_close",
        "last_close",
        "total_return_multiple",
        "total_return_pct",
        "start_close_used",
        "reference_return_multiple",
        "reference_return_pct",
        "data_points",
    ]
    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for h in int_cols:
        c = header_to_col.get(h)
        if not c:
            continue
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    for c in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(34, max_len + 2))

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Calculate adjusted-return metrics for all stocks with adjusted historical data")
    script_dir = Path(__file__).resolve().parent
    default_root = script_dir / "CANDLESTICKS" / "data" / "FETCH_CANDLESTICK_DATA"
    parser.add_argument("--root", default=str(default_root), help="Root folder containing industry/symbol/Historical/adjusted.json")
    parser.add_argument("--out-dir", default=str(script_dir), help="Output folder for CSV/XLSX")
    parser.add_argument("--coverage", type=float, default=0.80, help="Coverage threshold for global reference date (e.g., 0.80)")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    adjusted_files = sorted(root.glob("*/*/Historical/adjusted.json"))
    if not adjusted_files:
        raise FileNotFoundError(f"No adjusted.json files found under: {root}")

    series_list: list[StockSeries] = []
    for p in adjusted_files:
        s = parse_adjusted_file(p)
        if s is not None and len(s.dates) >= 2:
            series_list.append(s)

    if not series_list:
        raise RuntimeError("No valid stock series parsed")

    first_dates = [s.dates[0] for s in series_list]
    reference_date = find_reference_date(first_dates, coverage=args.coverage)

    rows = build_report_rows(series_list, reference_date)
    df = pd.DataFrame(rows)

    # Round all numeric outputs to no decimals (as requested).
    numeric_cols = [
        "first_close",
        "last_close",
        "total_return_multiple",
        "total_return_pct",
        "start_close_used",
        "reference_return_multiple",
        "reference_return_pct",
        "data_points",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(0).astype("Int64")

    csv_path = out_dir / "all_stocks_adjusted_returns.csv"
    xlsx_path = out_dir / "all_stocks_adjusted_returns.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="returns", index=False)
        ws = writer.book["returns"]
        from openpyxl.styles import PatternFill
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")

    autosize_excel(xlsx_path)

    print(f"Stocks parsed: {len(series_list)}")
    print(f"Global reference date ({args.coverage:.0%} coverage): {reference_date}")
    print(f"CSV:  {csv_path}")
    print(f"XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
